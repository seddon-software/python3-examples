from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

wb = Workbook()
ws = wb.active
dest_filename = 'empty_book.xlsx'
ws.title = "range names"

for row in range(5, 15):
    ws.append(range(25, 60, 3))

for row in range(20, 30):
    for col in range(1, 10):
        _ = ws.cell(column=col, row=row, value="--{}--".format(get_column_letter(col)))

wb.save('data/writingToCells.xlsx')