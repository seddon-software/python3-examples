from openpyxl import Workbook
wb = Workbook()

ws1 = wb.active  # get the first (and only) worksheet
ws2 = wb.create_sheet("Sheet 2")     # insert at the end (default)
ws0 = wb.create_sheet("Sheet 0", 0)  # insert at first position
ws1.title = "Sheet 1"
print wb.get_sheet_names()
wb.save('data/empty.xlsx')